#!/usr/bin/python
# -*- coding: UTF-8 -*-

import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import hashlib
import base64
import arrow
import re
from typing import List
import time
from datetime import datetime, timedelta
from lib import itchat
from lib.itchat.content import *
from channel.chat_message import ChatMessage
from croniter import croniter
import threading
import logging

# 获取logger
logger = logging.getLogger(__name__)

try:
    from channel.wechatnt.ntchat_channel import wechatnt
except Exception as e:
    print(f"未安装ntchat: {e}")

try:
    from channel.wework.run import wework
except Exception as e:
    print(f"未安装wework: {e}")

class ExcelTool(object):
    __file_name = "timeTask.xlsx"
    __sheet_name = "定时任务"
    __history_sheet_name = "历史任务"
    __dir_name = "taskFile"
    
    # 新建工作簿
    def create_excel(self, file_name: str = __file_name, sheet_name=__sheet_name, history_sheet_name=__history_sheet_name):
        # 文件路径
        workbook_file_path = self.get_file_path(file_name)

        # 创建Excel
        if not os.path.exists(workbook_file_path):
            wb = Workbook()
            column_list_first = ['A', 'B', 'C', 'D', 'L']
            width_value_first = 20
            column_list_two = ['E', 'F', 'H', 'J']
            width_value_two = 40
            column_list_three = ['G', 'I', 'K']
            width_value_three = 70
            width_value_four = 600
            
            # 设置日期格式
            date_format = NamedStyle(name='date_format')
            date_format.number_format = 'YYYY-MM-DD'

            #sheet1
            ws = wb.create_sheet(sheet_name, 0)
            # 类型处理
            for column in ws.columns:
                #日期格式
                if column == "D":
                    for cell in column:
                        cell.style = date_format
                #字符串        
                else:
                    for cell in column:
                        cell.number_format = '@'
            
            #宽度处理 
            for column in column_list_first:
                ws.column_dimensions[column].width = width_value_first
            for column in column_list_two:
                ws.column_dimensions[column].width = width_value_two
            for column in column_list_three:
                ws.column_dimensions[column].width = width_value_three
            ws.column_dimensions["M"].width = width_value_four 
              
            #sheet2
            ws1 = wb.create_sheet(history_sheet_name, 1)
            # 类型处理 - 设置为字符串
            for column in ws1.columns:
                for cell in column:
                    cell.number_format = '@'
                    
            #宽度处理        
            for column in column_list_first:
                ws1.column_dimensions[column].width = width_value_first
            for column in column_list_two:
                ws1.column_dimensions[column].width = width_value_two
            ws1.column_dimensions["M"].width = width_value_three     
                    
            wb.save(workbook_file_path)
            logger.info("定时Excel创建成功，文件路径为：{}".format(workbook_file_path))
            
        else:
            wb = load_workbook(workbook_file_path)
            if not history_sheet_name in wb.sheetnames:
                wb.create_sheet(history_sheet_name, 1)
                wb.save(workbook_file_path)
                logger.debug(f"创建sheet: {history_sheet_name}")
            else:
                logger.debug("timeTask文件已存在, 无需创建")
                

    # 读取内容,返回元组列表
    def readExcel(self, file_name=__file_name, sheet_name=__sheet_name):
        # 文件路径
        workbook_file_path = self.get_file_path(file_name)
        
        # 文件存在
        if os.path.exists(workbook_file_path):
            wb = load_workbook(workbook_file_path)
            ws = wb[sheet_name]
            data = list(ws.values)
            #print(data)
            if data is None or len(data) == 0:
                logger.debug("[timeTask] 数据库timeTask任务列表数据为空")
                
            return data
        else:
            print("timeTask文件不存在, 读取数据为空")
            self.create_excel()
            return []
        
    # 将历史任务迁移指历史Sheet
    def moveTasksToHistoryExcel(self, tasks, file_name=__file_name, sheet_name=__sheet_name, history_sheet_name=__history_sheet_name):
        # 文件路径
        workbook_file_path = self.get_file_path(file_name)
        
        # 文件存在
        if os.path.exists(workbook_file_path):
            wb = load_workbook(workbook_file_path)
            ws = wb[sheet_name]
            data = list(ws.values)
            
            #需要删除的坐标
            rows_to_delete = []
            #遍历任务列表
            for i, item in enumerate(data):
                 #任务ID
                 taskId = item[0]
                 for _, hisItem in enumerate(tasks):
                    #历史任务ID
                    his_taskId = hisItem[0]
                    if taskId == his_taskId:
                        rows_to_delete.append(i + 1)
            
            #排序坐标
            sorted_rows_to_delete = sorted(rows_to_delete, reverse=True)
                        
            #遍历任务列表
            for dx in sorted_rows_to_delete:
                #移除
                ws.delete_rows(dx)
                
            #保存            
            wb.save(workbook_file_path)
            
            hisIds = []
            #添加历史列表
            for _, t in enumerate(tasks):
                his_taskId = t[0]
                hisIds.append(his_taskId)
                self.addItemToExcel(t, file_name, history_sheet_name)     
                
            print(f"将任务Sheet({sheet_name})中的 过期任务 迁移指 -> 历史Sheet({history_sheet_name}) 完毕~ \n 迁移的任务ID为：{hisIds}")            
            
            #返回最新数据
            return self.readExcel()  
        else:
            print("timeTask文件不存在, 数据为空")
            self.create_excel()
            return []

    # 写入列表，返回元组列表
    def addItemToExcel(self, item, file_name=__file_name, sheet_name=__sheet_name):
        # 文件路径
        workbook_file_path = self.get_file_path(file_name)
        
        # 如果文件存在,就执行
        if os.path.exists(workbook_file_path):
            wb = load_workbook(workbook_file_path)
            ws = wb[sheet_name]
            ws.append(item)
            wb.save(workbook_file_path)
            
            # 列表
            data = list(ws.values)
            #print(data)
            return data
        else:
            print("timeTask文件不存在, 添加数据失败")
            self.create_excel()
            return []
        
        
    # 写入数据
    def write_columnValue_withTaskId_toExcel(self, taskId, column: int, columnValue: str,  file_name=__file_name, sheet_name=__sheet_name):
        #读取数据
        data = self.readExcel(file_name, sheet_name)
        if len(data) > 0:
            # 表格对象
            workbook_file_path = self.get_file_path(file_name)
            wb = load_workbook(workbook_file_path)
            ws = wb[sheet_name]
            isExist = False
            taskContent = None
            #遍历
            for index, hisItem in enumerate(data):
                model = TimeTaskModel(hisItem, None, False)
                #ID是否相同
                if model.taskId == taskId:
                    #置为已消费：即0
                    ws.cell(index + 1, column).value = columnValue
                    isExist = True
                    taskContent = model
                    
            if isExist: 
                #保存
                wb.save(workbook_file_path)
            
            return isExist, taskContent
        else:
            print("timeTask文件无数据, 消费数据失败")
            return False, None
    
    
    #获取文件路径      
    def get_file_path(self, file_name=__file_name):
        # 文件路径
        current_file = os.path.abspath(__file__)
        current_dir = os.path.dirname(current_file)
        workbook_file_path = current_dir + "/" + self.__dir_name + "/" + file_name
        
        # 工作簿当前目录
        workbook_dir_path = os.path.dirname(workbook_file_path)
        # 创建目录
        if not os.path.exists(workbook_dir_path):
            # 创建工作簿路径,makedirs可以创建级联路径
            os.makedirs(workbook_dir_path)
            
        return workbook_file_path
        
    #更新用户ID  
    def update_userId(self, file_name=__file_name, sheet_name=__sheet_name):
        #是否重新登录了
        datas = self.readExcel(file_name, sheet_name)
        
        if len(datas) <= 0:
            return
            
        #模型数组
        tempArray : List[TimeTaskModel] = []
        #原始数据
        for item in datas:
            model = TimeTaskModel(item, None, False)
            tempArray.append(model)
            
        #id字典数组：将相同目标人的ID聚合为一个数组
        idsDic = {}
        groupIdsDic = {}
        for model in tempArray:
            #目标用户名称
            target_name = model.other_user_nickname
            #群聊
            if model.isGroup:
                if not target_name in groupIdsDic.keys():
                    groupIdsDic[target_name] = [model]
                else:
                    arr1 = groupIdsDic[target_name]
                    arr1.append(model)
                    groupIdsDic[target_name] = list(arr1) 
            else:
                #好友
                if not target_name in idsDic.keys():
                    idsDic[target_name] = [model]
                else:
                    arr2 = idsDic[target_name]
                    arr2.append(model)
                    idsDic[target_name] = list(arr2)
        
        #待更新的ID数组
        if len(idsDic) <= 0:
            return
        
        #原始ID ：新ID
        oldAndNewIDDic = self.getNewId(idsDic, groupIdsDic)
        if len(oldAndNewIDDic) <= 0:
            return
            
        #更新列表数据
        workbook_file_path = self.get_file_path(file_name)
        wb = load_workbook(workbook_file_path)
        ws = wb[sheet_name]
        excel_data = list(ws.values)
        #机器人ID
        robot_user_id = itchat.instance.storageClass.userName
        #遍历任务列表 - 更新数据
        for index, item in enumerate(excel_data):
            model = TimeTaskModel(item, None, False)
            #目标用户ID
            oldId = model.other_user_id
            newId = oldAndNewIDDic.get(oldId)
            #找到了
            if newId is not None and len(newId) > 0:
                model.other_user_id = newId
                #更新ID
                #from
                ws.cell(index + 1, 7).value = newId
                #to
                ws.cell(index + 1, 9).value = robot_user_id
                #other
                ws.cell(index + 1, 11).value = newId
                #替换原始信息中的ID
                #旧的机器人ID
                old_robot_userId = model.toUser_id
                #原始消息体
                originStr = model.originMsg
                #替换旧的目标ID
                newString = originStr.replace(oldId, newId)
                #替换机器人ID
                newString = newString.replace(old_robot_userId, robot_user_id)
                ws.cell(index + 1, 13).value = newString
                #等待写入
                time.sleep(0.05)
                      
        #保存            
        wb.save(workbook_file_path)
        
            
            
    #获取新的用户ID  
    def getNewId(self, idsDic, groupIdsDic):
        oldAndNewIDDic = {}
        #好友  
        friends = []
        #群聊
        chatrooms = []
        
        #好友处理
        if len(idsDic) > 0:   
            #好友处理
            try:
                #获取好友列表
                friends = itchat.get_friends(update=True)[0:]
            except ZeroDivisionError:
                # 捕获并处理 ZeroDivisionError 异常
                print("好友列表, 错误发生")
            
            #获取好友 -（id组装 旧 ： 新）
            for friend in friends:
                #id
                userName = friend["UserName"]
                NickName = friend["NickName"]
                modelArray = idsDic.get(NickName)
                #找到了好友
                if modelArray is not None and len(modelArray) > 0:
                    model : TimeTaskModel = modelArray[0]
                    oldId = model.other_user_id
                    if oldId != userName:
                        oldAndNewIDDic[oldId] = userName    
         
        #群聊处理  
        if len(groupIdsDic) > 0:          
            #群聊处理       
            try:
                #群聊 （id组装 旧 ：新）   
                chatrooms = itchat.get_chatrooms()
            except ZeroDivisionError:
                # 捕获并处理 ZeroDivisionError 异常
                print("群聊列表, 错误发生")
            
            #获取群聊 - 旧 ： 新
            for chatroom in chatrooms:
                #id
                userName = chatroom["UserName"]
                NickName = chatroom["NickName"]
                modelArray = groupIdsDic.get(NickName)
                #找到了群聊
                if modelArray is not None and len(modelArray) > 0:
                    model : TimeTaskModel = modelArray[0]
                    oldId = model.other_user_id
                    if oldId != userName:
                        oldAndNewIDDic[oldId] = userName
                       
        return oldAndNewIDDic         
        

#task模型        
class TimeTaskModel:
    #Item数据排序
    #0：ID - 唯一ID (自动生成，无需填写)
    #1：是否可用 - 0/1，0=不可用，1=可用
    #2：时间信息 - 格式为：HH:mm:ss
    #3：轮询信息 - 格式为：每天、每周N、YYYY-MM-DD
    #4：消息内容 - 消息内容
    #5：fromUser - 来源user
    #6：fromUserID - 来源user ID
    #7：toUser - 发送给的user
    #8：toUser id - 来源user ID
    #9：other_user_nickname - Other名称
    #10：other_user_id - otehrID
    #11：isGroup - 0/1，是否群聊； 0=否，1=是
    #12：原始内容 - 原始的消息体
    #13：今天是否被消息 - 每天会在凌晨自动重置
    def __init__(self, item, msg:ChatMessage, isNeedFormat: bool, isNeedCalculateCron = True):
        
        self.isNeedCalculateCron = isNeedCalculateCron
        self.taskId = item[0]
        self.enable = item[1] == "1"
        
        #是否今日已被消费
        self.is_today_consumed = False
        
        #时间信息
        timeValue = item[2]
        tempTimeStr = ""
        if isinstance(timeValue, datetime):
            # 变量是 datetime.time 类型（Excel修改后，openpyxl会自动转换为该类型，本次做修正）
            tempTimeStr = timeValue.strftime("%H:%M:%S")
        elif isinstance(timeValue, str):
            tempTimeStr = timeValue
        else:
            # 其他类型
            logger.debug("不支持的时间格式")
        self.timeStr = tempTimeStr
        
        #日期
        dayValue = item[3]
        tempDayStr = ""
        if isinstance(dayValue, datetime):
            # 变量是 datetime.datetime 类型（Excel修改后，openpyxl会自动转换为该类型，本次做修正）
            tempDayStr = dayValue.strftime("%Y-%m-%d")
        elif isinstance(dayValue, str):
            tempDayStr = dayValue
        else:
            # 其他类型
            logger.debug("不支持的时间格式")
        self.circleTimeStr = tempDayStr
        
        #事件
        self.eventStr = item[4]
        
        #通过对象加载
        if msg is not None:
            self.fromUser = msg.from_user_nickname
            self.fromUser_id = msg.from_user_id
            self.toUser = msg.to_user_nickname
            self.toUser_id = msg.to_user_id
            self.other_user_nickname = msg.other_user_nickname
            self.other_user_id = msg.other_user_id
            self.isGroup = msg.is_group
            self.originMsg = str(msg)
        else:
            #通过Item加载
            self.fromUser = item[5]
            self.fromUser_id = item[6]
            self.toUser = item[7]
            self.toUser_id = item[8]
            self.other_user_nickname = item[9]
            self.other_user_id = item[10]
            self.isGroup = item[11] == "1"
            self.originMsg = item[12]
            if len(item) > 13:
                self.is_today_consumed = item[13] == "1" 
        
        #容错
        emptStr = ""
        self.fromUser = emptStr if self.fromUser is None else self.fromUser
        self.fromUser_id = emptStr if self.fromUser_id is None else self.fromUser_id
        self.toUser = emptStr if self.toUser is None else self.toUser
        self.toUser_id = emptStr if self.toUser_id is None else self.toUser_id
        self.other_user_nickname = emptStr if self.other_user_nickname is None else self.other_user_nickname
        self.other_user_id = emptStr if self.other_user_id is None else self.other_user_id
        self.isGroup = False if self.isGroup is None else self.isGroup
        self.originMsg = emptStr if self.originMsg is None else self.originMsg   
        
        #cron表达式
        self.cron_expression = self.get_cron_expression()
        
        #需要处理格式
        if isNeedFormat:
            #计算内容ID (使用不可变的内容计算，去除元素：enable 会变、originMsg中有时间戳)
            new_tuple = (self.timeStr, self.circleTimeStr, self.eventStr, self.fromUser, 
                         self.toUser, self.other_user_id, "1" if self.isGroup else "0")
            temp_content='_'.join(new_tuple)
            short_id = self.get_short_id(temp_content)
            print(f'消息体：{temp_content}， 唯一ID：{short_id}')
            self.taskId = short_id
            
            #周期、time
            #cron表达式
            if self.isCron_time():
                print("cron 表达式")
                
            else:
                #正常的周期、时间
                g_circle = self.get_cicleDay(self.circleTimeStr)
                g_time = self.get_time(self.timeStr)
                self.timeStr = g_time
                self.circleTimeStr = g_circle
                
        #今日消费态优化（默认程序在00:00会将消费态回写，但是若程序被kill,则下次启动的本地缓存未正确回写，此处需要容错）
        if self.is_today_consumed:
            if self.is_today() and (self.is_nowTime()[0] or self.is_featureTime()):
                self.is_today_consumed = False
                
        #数组为空
        self.cron_today_times = []
        
        #计算cron今天的时间点
        if self.isNeedCalculateCron and self.isCron_time() and self.enable:
            # 创建子线程
            t = threading.Thread(target=self.get_todayCron_times)
            t.setDaemon(True) 
            t.start() 
     
    #获取今天cron时间  
    def get_todayCron_times(self):
        if not self.enable:
              return
          
        self.cron_today_times = []
        #校验cron格式
        if self.isValid_Cron_time():
            # 获取当前时间（忽略秒数）
            current_time = arrow.now().replace(second=0, microsecond=0)
            # 创建一个 croniter 对象
            cron = croniter(self.cron_expression, current_time.datetime)
            next_time = cron.get_next(datetime)
            while next_time.date() == current_time.date():
                #记录时间（时：分）
                next_time_hour_minut = next_time.strftime('%H:%M')
                self.cron_today_times.append(next_time_hour_minut)
                next_time = cron.get_next(datetime)
            
            #打印满足今天的cron的时间点    
            print(f"cron表达式为：{self.cron_expression}, 满足今天的时间节点为：{self.cron_today_times}")
        
    #获取格式化后的Item
    def get_formatItem(self):
        temp_item = (self.taskId,
                "1" if self.enable else "0",
                self.timeStr,
                self.circleTimeStr,
                self.eventStr,
                self.fromUser,
                self.fromUser_id,
                self.toUser,
                self.toUser_id,
                self.other_user_nickname,
                self.other_user_id,
                "1" if self.isGroup else "0",
                self.originMsg,
                "1" if self.is_today_consumed else "0") 
        return temp_item
            
    #计算唯一ID        
    def get_short_id(self, string):
        # 使用 MD5 哈希算法计算字符串的哈希值
        hash_value = hashlib.md5(string.encode()).digest()
    
        # 将哈希值转换为一个 64 进制的短字符串
        short_id = base64.urlsafe_b64encode(hash_value)[:8].decode()
        return short_id
    
    
    #判断是否当前时间    
    def is_nowTime(self):
        """判断是否当前时间，返回(是否当前时间, 当前时间字符串)"""
        tempTimeStr = self.timeStr
        if not tempTimeStr:
            return False, ""
            
        # 如果只有时和分，补充秒
        if tempTimeStr.count(":") == 1:
           tempTimeStr = tempTimeStr + ":00"
        
        #cron   
        if self.isCron_time():
            current_time = arrow.now().replace(second=0, microsecond=0)
            current_time_str = current_time.format('HH:mm')
            return current_time_str in self.cron_today_times, current_time_str
        else:    
            # 获取当前时间和任务时间（保留秒）
            current_time = arrow.now().replace(microsecond=0)
            task_time = arrow.get(tempTimeStr, "HH:mm:ss")
            
            # 如果当前时间和任务时间在同一分钟内，立即执行
            if current_time.format('HH:mm') == task_time.format('HH:mm'):
                return True, current_time.format('HH:mm')
            
            # 如果当前时间比任务时间晚，且差距在1分钟内，认为是当前时间
            time_diff = (current_time - task_time).total_seconds() / 60
            is_now = 0 <= time_diff <= 1
            
            logger.debug(f"时间比较：当前时间={current_time.format('HH:mm:ss')}, 任务时间={task_time.format('HH:mm:ss')}, 时间差={time_diff}分钟, 是否执行={is_now}")
            
            return is_now, current_time.format('HH:mm')
    
    #判断是否未来时间    
    def is_featureTime(self):
        """判断是否未来时间"""
        tempTimeStr = self.timeStr
        if not tempTimeStr:
            return False
            
        if tempTimeStr.count(":") == 1:
           tempTimeStr = tempTimeStr + ":00"
        
        #cron   
        if self.isCron_time():
            return True 
        else:    
            #对比精准到分（忽略秒）
            current_time = arrow.now().replace(second=0, microsecond=0).time()
            task_time = arrow.get(tempTimeStr, "HH:mm:ss").replace(second=0, microsecond=0).time()
            tempValue = task_time > current_time
            return tempValue 
    
    #判断是否未来日期    
    def is_featureDay(self):
        """判断是否未来日期"""
        #cron   
        if self.isCron_time():
            return True
        
        else:     
            tempStr = self.circleTimeStr
            tempValue = "每天" in tempStr or "每周" in tempStr or "每星期" in tempStr  or "工作日" in tempStr
            #日期
            if self.is_valid_date(tempStr):
                tempValue = arrow.get(tempStr, 'YYYY-MM-DD').date() > arrow.now().date()
                
            return tempValue 
    
    #判断是否今天    
    def is_today(self):
        """判断是否今天"""
        try:
            # cron表达式处理
            if self.isCron_time():
                return True 
            
            # 当前时间
            current_date = arrow.now().format('YYYY-MM-DD')
            # 轮询信息
            item_circle = self.circleTimeStr
            
            # 1. 首先检查是否为空
            if not item_circle or item_circle.strip() == "":
                return True
            
            # 2. 检查是否包含"每天"关键字
            if item_circle == "每天" or item_circle == "cycle_每天":
                return True
            
            # 3. 处理其他周期性任务标记
            if item_circle.startswith("cycle_"):
                cycle_type = item_circle.replace("cycle_", "")
                
                if "每周" in cycle_type or "每星期" in cycle_type:
                    return self.is_today_weekday(cycle_type)
                    
                elif cycle_type == "工作日":
                    # 判断星期几（0-6，0是周一）
                    weekday = arrow.now().weekday()
                    # 判断是否是工作日（周一到周五）
                    return weekday < 5
                    
                return False
                
            # 4. 处理每周X格式（不带cycle_前缀）
            if "每周" in item_circle or "每星期" in item_circle:
                return self.is_today_weekday(item_circle)
                
            # 5. 处理工作日（不带cycle_前缀）
            if item_circle == "工作日":
                weekday = arrow.now().weekday()
                return weekday < 5
                
            # 6. 处理具体日期格式
            if self.is_valid_date(item_circle):
                return item_circle == current_date
                
            return False
            
        except Exception as e:
            logger.error(f"任务 {self.taskId} 检查日期时发生错误: {str(e)}")
            return False
    
    #判断是否今天的星期数    
    def is_today_weekday(self, weekday_str):
        """判断是否今天的星期数"""
        # 将中文数字转换为阿拉伯数字
        weekday_dict = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '日': 7, '天': 7}
        weekday_num = weekday_dict.get(weekday_str[-1])
        if weekday_num is None:
            return False
        
        # 判断今天是否是指定的星期几
        today = arrow.now()
        tempValue = today.weekday() == weekday_num - 1   
        return tempValue   
        
    #判断日期格式是否正确    
    def is_valid_date(self, date_string):
        """判断日期格式是否正确"""
        if not date_string:
            return False
        pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
        match = pattern.match(date_string)
        return match is not None
    
    #获取周期
    def get_cicleDay(self, circleStr):
        """获取格式化的日期"""
        if not circleStr:
            return ""
            
        g_circle = ""
        pattern1 = r'^\d{4}-\d{2}-\d{2}$'
        pattern2 = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$'
        pattern3 = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'
        
        try:
            # 如果是cron表达式，直接返回
            if circleStr.startswith("cron["):
                return circleStr
            
            # 处理周期性日期
            if circleStr == "每天" or circleStr == "cycle_每天":
                return "每天"
            
            if circleStr in ["每周", "工作日"] or circleStr in ["cycle_每周", "cycle_工作日"]:
                return circleStr.replace("cycle_", "")
            
            # 处理每周X
            if any(circleStr.endswith(x) for x in ["一", "二", "三", "四", "五", "六", "日", "天"]):
                if circleStr.startswith("cycle_"):
                    circleStr = circleStr.replace("cycle_", "")
                if circleStr.startswith("每周") or circleStr.startswith("每星期"):
                    return circleStr
            
            # 处理中文日期
            if circleStr in ['今天', '明天', '后天']:
                today = arrow.now('local')
                if circleStr == '今天':
                    g_circle = today.format('YYYY-MM-DD')
                elif circleStr == '明天':
                    g_circle = today.shift(days=1).format('YYYY-MM-DD')
                elif circleStr == '后天':
                    g_circle = today.shift(days=2).format('YYYY-MM-DD')
                return g_circle
            
            # 尝试解析标准日期格式
            if re.match(pattern1, circleStr):
                # 如果只有日期，添加时间
                g_circle = f"{circleStr} 00:00:00"
            elif re.match(pattern2, circleStr):
                # 如果有日期和时分，添加秒
                g_circle = f"{circleStr}:00"
            elif re.match(pattern3, circleStr):
                # 如果格式完整，直接使用
                g_circle = circleStr
            else:
                # 尝试智能解析日期
                try:
                    parsed_date = arrow.get(circleStr)
                    g_circle = parsed_date.format('YYYY-MM-DD HH:mm:ss')
                except Exception as e:
                    print(f"智能解析日期失败: {str(e)}")
                    return ""
                    
            logger.debug(f"转换日期: {circleStr} -> {g_circle}")
            return g_circle
            
        except Exception as e:
            logger.error(f"日期转换错误: {str(e)}")
            return ""
    
    def get_time(self, timeStr):
        """获取格式化的时间"""
        if not timeStr:
            return ""
            
        g_time = ""
        pattern1 = r'^\d{2}:\d{2}:\d{2}$'
        pattern2 = r'^\d{2}:\d{2}$'
        
        try:
            # 如果是cron表达式，直接返回
            if timeStr.startswith("cron["):
                return timeStr
                
            # 尝试解析标准时间格式
            if re.match(pattern1, timeStr):
                # 如果格式完整，直接使用
                g_time = timeStr
            elif re.match(pattern2, timeStr):
                # 如果只有时分，添加秒
                g_time = f"{timeStr}:00"
            else:
                # 处理中文时间描述
                try:
                    # 预处理时间字符串
                    content = timeStr.replace("早上", "").replace("上午", "").replace("中午", "").replace("下午", "").replace("晚上", "")
                    content = content.replace("点", ":").replace("分", ":").replace("秒", "")
                    
                    # 中文数字映射
                    digits = {
                        '零': 0, '一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10,
                        '十一': 11, '十二': 12, '十三': 13, '十四': 14, '十五': 15, '十六': 16, '十七': 17, '十八': 18, '十九': 19, '二十': 20,
                        '二十一': 21, '二十二': 22, '二十三': 23, '二十四': 24
                    }
                    
                    # 分解时间部分
                    parts = content.split(":")
                    hour = "0"
                    minute = "0"
                    second = "0"
                    
                    # 处理小时
                    if len(parts) > 0 and parts[0]:
                        if parts[0] in digits:
                            hour = str(digits[parts[0]])
                        else:
                            try:
                                hour = str(int(parts[0]))
                            except:
                                print(f"无法解析小时: {parts[0]}")
                                return ""
                                
                    # 处理分钟
                    if len(parts) > 1 and parts[1]:
                        if parts[1] in digits:
                            minute = str(digits[parts[1]])
                        else:
                            try:
                                minute = str(int(parts[1]))
                            except:
                                minute = "0"
                                
                    # 处理秒
                    if len(parts) > 2 and parts[2]:
                        if parts[2] in digits:
                            second = str(digits[parts[2]])
                        else:
                            try:
                                second = str(int(parts[2]))
                            except:
                                second = "0"
                                
                    # 处理时间段
                    if "中午" in timeStr:
                        if int(hour) < 12:
                            hour = "12"
                    elif "下午" in timeStr or "晚上" in timeStr:
                        if int(hour) < 12:
                            hour = str(int(hour) + 12)
                            
                    # 格式化时间
                    hour = f"0{hour}" if int(hour) < 10 else hour
                    minute = f"0{minute}" if int(minute) < 10 else minute
                    second = f"0{second}" if int(second) < 10 else second
                    
                    g_time = f"{hour}:{minute}:{second}"
                    
                except Exception as e:
                    print(f"解析中文时间失败: {str(e)}")
                    return ""
                    
            logger.debug(f"转换时间: {timeStr} -> {g_time}")
            
            # 验证最终时间格式
            if re.match(pattern1, g_time):
                return g_time
            return ""
            
        except Exception as e:
            logger.error(f"时间转换错误: {str(e)}")
            return ""
    
    #是否 cron表达式
    def isCron_time(self):
        tempValue = self.circleTimeStr.startswith("cron[")
        return tempValue
    
    #是否正确的cron格式
    def isValid_Cron_time(self):
        tempValue = croniter.is_valid(self.cron_expression)
        return tempValue
    
    #获取 cron表达式
    def get_cron_expression(self):
        tempValue = self.timeStr
        tempValue = tempValue.replace("cron[", "")
        tempValue = tempValue.replace("Cron[", "")
        tempValue = tempValue.replace("]", "")
        return tempValue
    
    #是否 私聊制定群任务
    def isPerson_makeGrop(self):
        tempValue = self.eventStr.endswith("]")
        tempValue1 = "group[" in self.eventStr or "Group[" in self.eventStr
        return tempValue and tempValue1
    
    #获取私聊制定群任务的群Title、事件
    def get_Persion_makeGropTitle_eventStr(self):
        index = -1
        targetStr = self.eventStr
        if "group[" in targetStr:
            index = targetStr.index("group[")
        elif "Group[" in targetStr:
            index = targetStr.index("Group[")
        if index < 0:
              return "", targetStr
          
        substring_event = targetStr[:index].strip()
        substring_groupTitle = targetStr[index + 6:]
        substring_groupTitle = substring_groupTitle.replace("]", "").strip()
        return substring_event, substring_groupTitle
    
    #通过 群Title 获取群ID或群名称
    def get_gropID_withGroupTitle(self, groupTitle, channel_name):
        """通过群标题获取群ID或群名称"""
        if len(groupTitle) <= 0:
              return ""
              
        #itchat - 返回群ID
        if channel_name == "wx":
            try:
                #群聊处理       
                chatrooms = itchat.get_chatrooms(update=True)  # 强制更新群列表
                if not chatrooms:
                    # 如果获取失败，等待1秒后重试一次
                    time.sleep(1)
                    chatrooms = itchat.get_chatrooms(update=True)
                
                logger.info(f"[{channel_name}] 开始查找群【{groupTitle}】")
                #获取群聊
                for chatroom in chatrooms:
                    NickName = chatroom["NickName"]
                    # 使用大小写不敏感的精确匹配
                    if NickName.lower() == groupTitle.lower():
                        logger.info(f"[{channel_name}] 找到匹配的群：【{NickName}】，群ID：{chatroom['UserName']}")
                        return chatroom["UserName"]  # 返回群ID而不是群名称
                
                logger.error(f"[{channel_name}] 未找到群【{groupTitle}】，当前共有 {len(chatrooms)} 个群")
                # 记录所有群名称，用于调试
                for chatroom in chatrooms:
                    logger.error(f"[{channel_name}] 可用的群：【{chatroom['NickName']}】")
                return ""
                    
            except Exception as e:
                logger.error(f"[{channel_name}] 通过群标题获取群ID时发生错误：{str(e)}")
                return ""

        elif channel_name == "ntchat":
            tempRoomId = ""
            try:
                #数据结构为字典数组
                rooms = wechatnt.get_rooms()
                
                if len(rooms) > 0:
                    #遍历
                    for item in rooms:
                        roomId = item.get("wxid")
                        nickname = item.get("nickname")
                        # 使用精确匹配
                        if nickname == groupTitle:
                            tempRoomId = roomId
                            break
                            
                return tempRoomId
                        
            except Exception as e:
                logger.error(f"[{channel_name}] 通过群标题获取群ID时发生错误：{str(e)}")
                return tempRoomId

        elif channel_name == "wework":
            tempRoomId = ""
            try:
                # 数据结构为字典数组
                rooms = wework.get_rooms().get("room_list")
                
                if len(rooms) > 0:
                    # 遍历
                    for item in rooms:
                        roomId = item.get("conversation_id")
                        nickname = item.get("nickname")
                        # 使用精确匹配
                        if nickname == groupTitle:
                            tempRoomId = roomId
                            break
                            
                return tempRoomId
                        
            except Exception as e:
                logger.error(f"[{channel_name}] 通过群标题获取群ID时发生错误：{str(e)}")
                return tempRoomId

        else:
            logger.error(f"[{channel_name}] 不支持通过群标题获取群ID，当前channel：{channel_name}")
            return ""

class CleanFiles:
    def __init__(self, save_path):
        self.save_path = save_path

    def clean_expired_files(self, days=3):
        """清理过期文件"""
        try:
            # 使用更灵活的时间格式解析
            current_time = datetime.now()
            expire_time = current_time - timedelta(days=days)
            
            # 遍历目录
            for root, dirs, files in os.walk(self.save_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    try:
                        # 获取文件修改时间
                        file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                        if file_time < expire_time:
                            try:
                                os.remove(file_path)
                                logger.info(f"已删除过期文件: {file_path}")
                            except Exception as e:
                                logger.error(f"删除文件失败 {file_path}: {str(e)}")
                    except Exception as e:
                        logger.error(f"获取文件时间失败 {file_path}: {str(e)}")
        except Exception as e:
            logger.error(f"清理过期文件出错: {str(e)}")
